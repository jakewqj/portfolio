"""
Import PR coverage from the master xlsx into src/data/coverage.json.

Run:
    python scripts/import-coverage.py
    python scripts/import-coverage.py --xlsx "D:/Work/Raw/PR coverages 20260507 - 副本.xlsx"

Outputs:
    src/data/coverage.json       site data
    .coverage-import.log         issues to triage (unrecognized outlets, parse errors)
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import yaml

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = Path(r"D:/Work/Raw/PR coverages 20260507 - 副本.xlsx")
OUTLETS_YAML = REPO_ROOT / "src" / "data" / "outlets.yaml"
OUT_JSON = REPO_ROOT / "src" / "data" / "coverage.json"
LOG_FILE = REPO_ROOT / ".coverage-import.log"

# Per-sheet column map (1-based, matches openpyxl).
# `reach_alt` covers an alternate column where reach was sometimes typed by mistake.
# In the 2023 sheet, ~74 rows have the MUV value in col 11 (RANK) instead of col 10
# (Monthly Visitors); falling back recovers them without re-touching the xlsx.
#
# 2023 sheet has no separate '内容形式' column; the legacy '形式' col 7 holds the
# content type (通稿 / 测评 / 榜单 / 约稿 / etc.). Map it as content_type and let
# format default to 'online' (all rows pass the http-URL filter).
SHEETS = {
    "2023": {"date": 1, "outlet_category": 2, "product": 3, "outlet": 4, "url": 5,
             "content_type": 7, "reach": 10, "reach_alt": 11},
    "2024 Coverage": {"date": 1, "region": 2, "product": 4, "outlet": 5,
                      "outlet_category": 6, "url": 7, "format": 8, "content_type": 9,
                      "reach": 10},
    "2025": {"date": 1, "region": 2, "product": 4, "outlet": 5,
             "outlet_category": 6, "url": 7, "format": 8, "content_type": 9,
             "reach": 10},
    "2026": {"date": 1, "region": 2, "product": 4, "outlet": 5,
             "outlet_category": 6, "url": 7, "format": 8, "content_type": 9,
             "reach": 10},
}

OUTLET_CATEGORY_MAP = {
    "乐器垂类": "instrument-vertical",
    "音乐文化": "music-culture",
    "大众媒体": "mass-media",
    "大众文化": "mass-media",
    "其他": "other",
    "Newswire": "newswire",
}

FORMAT_MAP = {"线上": "online", "印刷": "print"}

# Legacy 2023 '形式' values → 2024+ content_type taxonomy.
CONTENT_TYPE_MAP = {
    "通稿": "News",
    "通稿宣发": "News",
    "通稿联系报道": "News",
    "约稿": "Feature",
    "测评": "Review",
    "文字测评": "Review",
    "Youtube视频测评": "Review",
    "榜单": "Round-up",
    "推荐": "Round-up",
    "媒体购买": "Ad",
    "广告": "Ad",
    "社媒": "Social",
    "媒体转发": "Social",
    "媒体转载": "Social",
    "内容合作": "Feature",
}


def slugify(text: str) -> str:
    """Lowercase, replace non-alphanumeric with hyphens. Preserves CJK."""
    text = text.strip().lower()
    text = re.sub(r"[\s/_]+", "-", text)
    text = re.sub(r"[^a-z0-9\-一-鿿]", "", text)
    return text.strip("-") or "unknown"


def parse_date(value, sheet_year: int):
    """Coerce a cell's date-ish value into ISO YYYY-MM-DD. Falls back to year-only."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if not value:
        return f"{sheet_year}-01-01"
    s = str(value).strip()
    # "Q1", "Q3" → midpoint of quarter
    qmatch = re.match(r"^Q([1-4])$", s, re.IGNORECASE)
    if qmatch:
        q = int(qmatch.group(1))
        month = (q - 1) * 3 + 2
        return f"{sheet_year}-{month:02d}-15"
    # "January 30, 2023" / "Sat, May 6, 2023" / "2025-01-01 00:00:00"
    for fmt in ("%B %d, %Y", "%a, %b %d, %Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%b %d, %Y", "%b, %Y", "%B, %Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    # Year fallback.
    return f"{sheet_year}-01-01"


def parse_reach(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value) if value > 0 else None
    s = str(value).strip()
    if not s or s in {"数据不够", "待定", "/"}:
        return None
    s = s.replace(",", "").replace(" ", "")
    # "1.7M" / "5K"
    suffix = {"K": 1_000, "M": 1_000_000, "B": 1_000_000_000}
    if s and s[-1].upper() in suffix:
        try:
            return int(float(s[:-1]) * suffix[s[-1].upper()])
        except ValueError:
            return None
    try:
        return int(float(s))
    except ValueError:
        return None


class OutletResolver:
    def __init__(self, yaml_path: Path):
        with yaml_path.open(encoding="utf-8") as f:
            cfg = yaml.safe_load(f)
        self.outlets = cfg["outlets"]
        self.campaigns = cfg["campaigns"]
        self.domain_region = cfg.get("domain_region_fallback", {})
        # alias → canonical outlet
        self.alias_index: dict[str, dict] = {}
        # domain → canonical outlet (used as fallback when xlsx outlet cell is blank)
        self.domain_index: dict[str, dict] = {}
        for o in self.outlets:
            keys = {o["name"]}
            keys.update(o.get("aliases", []))
            for k in keys:
                self.alias_index[k.lower().strip()] = o
            if o.get("domain"):
                self.domain_index[o["domain"].lower()] = o
        # campaign matchers
        self.campaign_matchers = []
        for c in self.campaigns:
            patterns = [re.compile(p, re.IGNORECASE) for p in c["patterns"]]
            self.campaign_matchers.append((c["slug"], patterns))

    def resolve(self, raw_outlet: str) -> dict | None:
        if not raw_outlet:
            return None
        return self.alias_index.get(raw_outlet.lower().strip())

    def resolve_by_url(self, url: str) -> dict | None:
        try:
            host = urlparse(url).netloc.lower().lstrip("www.")
        except Exception:
            return None
        # exact match, then suffix match (e.g. fr.audiofanzine.com → audiofanzine.com)
        if host in self.domain_index:
            return self.domain_index[host]
        for dom, meta in self.domain_index.items():
            if host.endswith("." + dom) or host == dom:
                return meta
        return None

    def match_campaign(self, product: str) -> str | None:
        if not product:
            return None
        for slug, patterns in self.campaign_matchers:
            for pat in patterns:
                if pat.search(product):
                    return slug
        return None

    def region_from_domain(self, url: str) -> str | None:
        try:
            host = urlparse(url).netloc.lower()
        except Exception:
            return None
        for suffix, region in self.domain_region.items():
            if host.endswith(suffix):
                return region
        return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--out", type=Path, default=OUT_JSON)
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"xlsx not found: {args.xlsx}")

    resolver = OutletResolver(OUTLETS_YAML)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    clippings: list[dict] = []
    seen_ids: dict[str, int] = {}
    unknown_outlets: Counter[str] = Counter()
    skipped = 0

    for sheet_name, cmap in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_year = int(re.match(r"\d{4}", sheet_name).group())
        for r in range(2, ws.max_row + 1):
            url = ws.cell(row=r, column=cmap["url"]).value
            if not (isinstance(url, str) and url.startswith("http")):
                skipped += 1
                continue

            raw_outlet = ws.cell(row=r, column=cmap["outlet"]).value
            raw_outlet = str(raw_outlet).strip() if raw_outlet else ""
            outlet_meta = resolver.resolve(raw_outlet)

            # If outlet cell is blank (or unrecognized), try inferring from URL domain.
            if not outlet_meta:
                inferred = resolver.resolve_by_url(url)
                if inferred:
                    outlet_meta = inferred
                    if not raw_outlet:
                        raw_outlet = inferred["name"]

            if not outlet_meta and raw_outlet:
                unknown_outlets[raw_outlet] += 1

            outlet_name = outlet_meta["name"] if outlet_meta else raw_outlet or "Unknown"
            outlet_tier = outlet_meta.get("tier") if outlet_meta else 3
            outlet_logo = outlet_meta.get("logo") if outlet_meta else None
            default_muv = outlet_meta.get("default_muv") if outlet_meta else None

            date_iso = parse_date(ws.cell(row=r, column=cmap["date"]).value, sheet_year)
            reach = parse_reach(ws.cell(row=r, column=cmap["reach"]).value)
            if reach is None and "reach_alt" in cmap:
                reach = parse_reach(ws.cell(row=r, column=cmap["reach_alt"]).value)
            if reach is None:
                reach = default_muv

            product_raw = ws.cell(row=r, column=cmap["product"]).value
            product_str = str(product_raw).strip() if product_raw else ""
            campaign_slug = resolver.match_campaign(product_str)

            region = None
            if "region" in cmap:
                region_raw = ws.cell(row=r, column=cmap["region"]).value
                region = str(region_raw).strip() if region_raw else None
            if not region:
                region = (outlet_meta.get("region") if outlet_meta else None) or resolver.region_from_domain(url)

            outlet_category = None
            if "outlet_category" in cmap:
                cat_raw = ws.cell(row=r, column=cmap["outlet_category"]).value
                outlet_category = OUTLET_CATEGORY_MAP.get(str(cat_raw).strip()) if cat_raw else None
            if not outlet_category and outlet_meta:
                outlet_category = outlet_meta.get("category")

            content_type = None
            if "content_type" in cmap:
                ct_raw = ws.cell(row=r, column=cmap["content_type"]).value
                if ct_raw:
                    ct_str = str(ct_raw).strip()
                    content_type = CONTENT_TYPE_MAP.get(ct_str, ct_str)

            fmt = None
            if "format" in cmap:
                fmt_raw = ws.cell(row=r, column=cmap["format"]).value
                fmt = FORMAT_MAP.get(str(fmt_raw).strip(), str(fmt_raw).strip()) if fmt_raw else None
            # 2023 sheet has no media-format column; everything that survived the
            # http-URL filter is web coverage.
            if fmt is None and "format" not in cmap:
                fmt = "online"

            base_id = f"{date_iso}-{slugify(outlet_name)}"
            count = seen_ids.get(base_id, 0)
            seen_ids[base_id] = count + 1
            cid = base_id if count == 0 else f"{base_id}-{count + 1}"

            clippings.append({
                "id": cid,
                "date": date_iso,
                "year": sheet_year,
                "region": region,
                "outlet": outlet_name,
                "outlet_tier": outlet_tier,
                "outlet_category": outlet_category,
                "outlet_logo": outlet_logo,
                "product": product_str or None,
                "campaign": campaign_slug,
                "content_type": content_type,
                "format": fmt,
                "reach": reach,
                "url": url,
            })

    clippings.sort(key=lambda c: (c["date"], c["outlet"]), reverse=True)

    # Aggregates.
    total_reach = sum((c["reach"] or 0) for c in clippings)
    unique_outlets = len({c["outlet"] for c in clippings})
    regions = sorted({c["region"] for c in clippings if c["region"]})
    by_year = Counter(c["year"] for c in clippings)
    by_campaign = Counter(c["campaign"] for c in clippings if c["campaign"])

    payload = {
        "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source_xlsx": args.xlsx.name,
        "totals": {
            "clippings": len(clippings),
            "outlets": unique_outlets,
            "regions": regions,
            "cumulative_reach": total_reach,
            "by_year": dict(sorted(by_year.items())),
            "by_campaign": dict(by_campaign),
        },
        "clippings": clippings,
    }

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    log_lines = [
        f"# coverage import — {payload['generated_at']}",
        f"clippings written: {len(clippings)}",
        f"skipped (no http URL): {skipped}",
        f"cumulative reach: {total_reach:,}",
        f"unique outlets: {unique_outlets}",
        f"by year: {dict(sorted(by_year.items()))}",
        f"by campaign: {dict(by_campaign)}",
        "",
        f"## unrecognized outlets ({len(unknown_outlets)}):",
        *(f"  {name!r}: {count}" for name, count in unknown_outlets.most_common()),
    ]
    LOG_FILE.write_text("\n".join(log_lines), encoding="utf-8")
    print("\n".join(log_lines[:7]))
    print(f"\n→ {args.out.relative_to(REPO_ROOT)}")
    print(f"→ {LOG_FILE.relative_to(REPO_ROOT)} (review unrecognized outlets, add to outlets.yaml)")


if __name__ == "__main__":
    main()
